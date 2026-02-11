import csv
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def read_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        return list(reader)

def write_xlsx(rows, out_path):
    wb = Workbook()
    ws = wb.active
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for i, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

def main():
    csv_path = os.path.join('docs', 'WBS.csv')
    xlsx_path = os.path.join('docs', 'WBS.xlsx')
    if not os.path.exists(csv_path):
        print('CSV not found:', csv_path)
        return 1
    rows = read_csv(csv_path)
    write_xlsx(rows, xlsx_path)
    print('WBS.xlsx written to', xlsx_path)
    return 0

if __name__ == '__main__':
    import sys
    sys.exit(main())
